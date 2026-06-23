"""Export presale purchase accounts (address + transaction info) to XLSX."""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
SRC = OUTPUT_DIR / "pes-purchase-accounts-bsc-mainnet.json"


def export_xlsx(payload: dict, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    accounts = payload["accounts"]
    wb = Workbook()
    ws = wb.active
    ws.title = "预售认购地址交易"

    header_fill = PatternFill("solid", fgColor="0B1020")
    header_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")

    ws.append(["PES 预售认购用户地址与交易信息"])
    ws.append([f"预售合约: {payload.get('presale', '')}"])
    ws.append(
        [
            f"区块 #{payload.get('fromBlock', '')} - #{payload.get('toBlock', '')}  |  "
            f"用户数 {payload.get('uniqueBuyerCount', len(accounts))}  |  "
            f"事件数 {payload.get('eventCount', len(accounts))}"
        ]
    )
    ws.append([])

    headers = [
        "序号",
        "买家地址",
        "认购份数",
        "支付USDT",
        "获得PES",
        "交易次数",
        "首次认购时间(UTC)",
        "交易哈希",
        "BscScan链接",
    ]
    keys = [
        "index",
        "buyer",
        "packages",
        "paymentUSDT",
        "tokenPES",
        "txCount",
        "firstTime",
        "firstTx",
        "bscscanTx",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in accounts:
        enriched = {
            **row,
            "bscscanTx": f"https://bscscan.com/tx/{row.get('firstTx', '')}" if row.get("firstTx") else "",
        }
        ws.append([enriched.get(key, "") for key in keys])

    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 48)

    wb.save(path)


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else SRC
    payload = json.loads(src.read_text(encoding="utf-8"))
    path = OUTPUT_DIR / "pes-purchase-accounts-bsc-mainnet.xlsx"
    export_xlsx(payload, path)
    print(path)


if __name__ == "__main__":
    main()

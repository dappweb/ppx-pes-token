"""Export released user table to CSV and XLSX."""
from __future__ import annotations

import csv
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"


def latest_audit_json() -> Path:
    files = sorted(OUTPUT_DIR.glob("pes-user-distribution-audit-*.json"))
    if not files:
        raise FileNotFoundError("请先运行: node output/audit-user-distributions.js")
    return files[-1]


def export_csv(rows: list[dict], path: Path, summary: dict) -> None:
    fieldnames = [
        "index",
        "address",
        "claimedPES",
        "allocatedPES",
        "vestedPES",
        "claimablePES",
        "expectedAtElapsedPES",
        "status",
        "firstPurchaseTime",
        "firstPurchaseTx",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})
        writer.writerow({})
        writer.writerow({"index": "汇总", "address": f"用户数 {summary['users']}", "claimedPES": summary["totalClaimed"]})


def export_xlsx(rows: list[dict], path: Path, summary: dict, meta: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "已发放用户"

    headers = [
        "序号",
        "用户地址",
        "已释放(PES)",
        "归属登记(PES)",
        "当期应发(PES)",
        "可领取(PES)",
        "状态",
        "首次购买时间",
        "首次购买Tx",
    ]
    keys = [
        "index",
        "address",
        "claimedPES",
        "allocatedPES",
        "expectedAtElapsedPES",
        "claimablePES",
        "status",
        "firstPurchaseTime",
        "firstPurchaseTx",
    ]

    ws.append(["PES 释放合约已发放用户清单"])
    ws.append([f"归属合约: {meta['presale']}"])
    ws.append([f"数据区块: #{meta['blockNumber']}  |  释放总额: {summary['totalClaimed']} PES  |  用户数: {summary['users']}"])
    ws.append([])

    header_fill = PatternFill("solid", fgColor="0B1020")
    header_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(key, "") for key in keys])

    summary_row = len(rows) + 6
    ws.cell(row=summary_row, column=1, value="合计")
    ws.cell(row=summary_row, column=2, value=f"{summary['users']} 个地址")
    ws.cell(row=summary_row, column=3, value=float(summary["totalClaimed"]))

    widths = [8, 46, 14, 14, 14, 12, 12, 22, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    ws.freeze_panes = "A6"
    wb.save(path)


def main() -> None:
    audit_path = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_audit_json()
    data = json.loads(audit_path.read_text(encoding="utf-8"))
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    rows = sorted(data["accounts"], key=lambda item: int(item["index"]))
    total_claimed = sum(float(row["claimedPES"]) for row in rows)
    summary = {"users": len(rows), "totalClaimed": f"{total_claimed:.1f}"}
    meta = {
        "presale": data["contracts"]["presaleVesting"],
        "blockNumber": data["blockNumber"],
    }

    csv_path = OUTPUT_DIR / f"pes-released-users-{stamp}.csv"
    xlsx_path = OUTPUT_DIR / f"pes-released-users-{stamp}.xlsx"

    export_csv(rows, csv_path, summary)
    export_xlsx(rows, xlsx_path, summary, meta)

    print(str(csv_path))
    print(str(xlsx_path))


if __name__ == "__main__":
    main()

"""Export DEX pool PES purchase trades to CSV and XLSX."""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"


def latest_json() -> Path:
    files = sorted(OUTPUT_DIR.glob("pes-dex-pool-purchases-*.json"))
    if not files:
        raise FileNotFoundError("请先运行: node output/export_dex_pool_purchases.js")
    return files[-1]


def style_header_row(ws, col_count: int, fill, font) -> None:
    from openpyxl.styles import Alignment

    for col in range(1, col_count + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autosize_sheet(ws, max_width: int = 48) -> None:
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def export_csv_trades(trades: list[dict], path: Path, summary: dict) -> None:
    fields = [
        "index",
        "buyer",
        "recipient",
        "pesAmount",
        "usdtAmount",
        "priceUSDTPerPES",
        "timeUTC",
        "blockNumber",
        "txHash",
        "bscscanTx",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in trades:
            writer.writerow({key: row.get(key, "") for key in fields})
        writer.writerow({})
        writer.writerow(
            {
                "index": "汇总",
                "buyer": f"买家数 {summary['uniqueBuyers']}",
                "pesAmount": summary["totalPESPurchased"],
                "usdtAmount": summary["totalUSDTPaid"],
            }
        )


def export_csv_buyers(buyers: list[dict], path: Path) -> None:
    fields = [
        "index",
        "buyer",
        "tradeCount",
        "totalPES",
        "totalUSDT",
        "avgPriceUSDTPerPES",
        "firstTime",
        "lastTime",
        "firstTx",
        "lastTx",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in buyers:
            writer.writerow({key: row.get(key, "") for key in fields})


def export_csv_pair_txs(rows: list[dict], path: Path) -> None:
    fields = [
        "index",
        "type",
        "from",
        "to",
        "method",
        "blockNumber",
        "timeUTC",
        "gasUsed",
        "status",
        "txHash",
        "bscscanTx",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fields})


def append_table_sheet(
    wb,
    title: str,
    title_rows: list[list],
    headers: list[str],
    keys: list[str],
    rows: list[dict],
    header_fill,
    header_font,
) -> None:
    ws = wb.create_sheet(title)
    for row in title_rows:
        ws.append(row)
    if title_rows:
        ws.append([])
    ws.append(headers)
    style_header_row(ws, len(headers), header_fill, header_font)
    for row in rows:
        ws.append([row.get(key, "") for key in keys])
    autosize_sheet(ws)


def export_xlsx(payload: dict, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    summary = payload["summary"]
    trades = payload["trades"]
    buyers = payload["buyers"]
    pair_txs = payload.get("pairAddressTransactions", [])

    header_fill = PatternFill("solid", fgColor="0B1020")
    header_font = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")

    wb = Workbook()
    ws = wb.active
    ws.title = "逐笔买入"

    ws.append(["PES/USDT 底池买入记录（PancakeSwap V2）"])
    ws.append([f"交易对: {summary['pair']}"])
    ws.append([f"BscScan: {summary.get('bscscanPair', '')}"])
    ws.append(
        [
            f"区块 #{summary['fromBlock']} - #{summary['toBlock']}  |  "
            f"买入笔数 {summary['buyTrades']}  |  买家 {summary['uniqueBuyers']}  |  "
            f"合计 {summary['totalPESPurchased']} PES / {summary['totalUSDTPaid']} USDT"
        ]
    )
    if summary.get("note"):
        ws.append([summary["note"]])
    ws.append([])

    trade_headers = [
        "序号",
        "买家地址",
        "接收地址",
        "买入PES",
        "支付USDT",
        "单价(USDT/PES)",
        "时间(UTC)",
        "区块",
        "交易哈希",
        "BscScan链接",
    ]
    trade_keys = [
        "index",
        "buyer",
        "recipient",
        "pesAmount",
        "usdtAmount",
        "priceUSDTPerPES",
        "timeUTC",
        "blockNumber",
        "txHash",
        "bscscanTx",
    ]
    ws.append(trade_headers)
    style_header_row(ws, len(trade_headers), header_fill, header_font)
    for row in trades:
        ws.append([row.get(key, "") for key in trade_keys])
    autosize_sheet(ws)

    append_table_sheet(
        wb,
        "按买家汇总",
        [],
        [
            "序号",
            "买家地址",
            "买入次数",
            "合计PES",
            "合计USDT",
            "均价(USDT/PES)",
            "首次买入",
            "末次买入",
            "首笔Tx",
            "末笔Tx",
        ],
        [
            "index",
            "buyer",
            "tradeCount",
            "totalPES",
            "totalUSDT",
            "avgPriceUSDTPerPES",
            "firstTime",
            "lastTime",
            "firstTx",
            "lastTx",
        ],
        buyers,
        header_fill,
        header_font,
    )

    append_table_sheet(
        wb,
        "交易对地址交易",
        [
            [f"交易对合约: {summary['pair']}"],
            ["该地址收到的链上普通交易（含 Approve / Swap 等）"],
        ],
        [
            "序号",
            "类型",
            "发起地址",
            "目标地址",
            "方法",
            "区块",
            "时间(UTC)",
            "Gas消耗",
            "状态",
            "交易哈希",
            "BscScan链接",
        ],
        [
            "index",
            "type",
            "from",
            "to",
            "method",
            "blockNumber",
            "timeUTC",
            "gasUsed",
            "status",
            "txHash",
            "bscscanTx",
        ],
        pair_txs,
        header_fill,
        header_font,
    )

    wb.save(path)


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_json()
    payload = json.loads(src.read_text(encoding="utf-8"))
    date = src.stem.replace("pes-dex-pool-purchases-", "")

    trades_csv = OUTPUT_DIR / f"pes-dex-pool-purchases-{date}.csv"
    buyers_csv = OUTPUT_DIR / f"pes-dex-pool-buyers-{date}.csv"
    pair_txs_csv = OUTPUT_DIR / f"pes-dex-pool-address-txs-{date}.csv"
    xlsx = OUTPUT_DIR / f"pes-dex-pool-purchases-{date}.xlsx"

    export_csv_trades(payload["trades"], trades_csv, payload["summary"])
    export_csv_buyers(payload["buyers"], buyers_csv)
    export_csv_pair_txs(payload.get("pairAddressTransactions", []), pair_txs_csv)
    export_xlsx(payload, xlsx)

    print(trades_csv)
    print(buyers_csv)
    print(pair_txs_csv)
    print(xlsx)


if __name__ == "__main__":
    main()
